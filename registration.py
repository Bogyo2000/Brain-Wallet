from tkinter import *
from openpyxl import *
from tkinter import messagebox
import re
import os
import main
import words
import images


def reg():
    def prevpage():
        root.destroy()
        main.main()

    def regname_focus():
        register_name.focus_set()

    def regemail_focus():
        register_email.focus_set()

    def reglogintype_focus():
        register_logintype.focus_set()

    # def clear():
    #     register_name.delete(0, END)
    #     register_email.delete(0, END)

    def excel():
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 100
        sheet.column_dimensions['E'].width = 60

        sheet.cell(row=1, column=1).value = "Username"
        sheet.cell(row=1, column=2).value = "Email"
        sheet.cell(row=1, column=3).value = "Login type"
        sheet.cell(row=1, column=4).value = "Words hash"
        sheet.cell(row=1, column=5).value = "Image hash"

    def existingname():
        for row in sheet.iter_rows():
            if (row[0].value == register_name.get() or row[1].value == register_email.get()):
                return True
        return False

    def isValid(register_email):
        regex = r'([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+'
        if re.fullmatch(regex, register_email):
            return True
        return False

    def wordsorimages():
        if (register_name.get() == "" or
                register_email.get() == "" or var.get() != 2 and var.get() != 1):
            messagebox.showerror("Error", "All fields are required", parent=root)

        elif (existingname()):
            messagebox.showerror("Error", "Name or email is already existing", parent=root)

        elif not isValid(register_email.get()):
            messagebox.showerror("Error", "Invalid email", parent=root)

        elif register_name.get() != "" and register_email.get() != "" and var.get() == 1:
            current_row = sheet.max_row

            sheet.cell(row=current_row + 1, column=1).value = register_name.get()
            sheet.cell(row=current_row + 1, column=2).value = register_email.get()
            sheet.cell(row=current_row + 1, column=3).value = var.get()

            root.destroy()
            words.words()
            word_hash = words.key

            sheet.cell(row=current_row + 1, column=4).value = str(word_hash)
            currentdir = os.getcwd()
            file = os.path.join(currentdir, "excel.xlsx")
            wb.save(file)
            wb.close()
            main.main()

        else:
            current_row = sheet.max_row

            sheet.cell(row=current_row + 1, column=1).value = register_name.get()
            sheet.cell(row=current_row + 1, column=2).value = register_email.get()
            sheet.cell(row=current_row + 1, column=3).value = var.get()

            root.destroy()
            images.images()
            img_hash = images.img_hash

            sheet.cell(row=current_row + 1, column=5).value = img_hash
            currentdir = os.getcwd()
            file = os.path.join(currentdir, "excel.xlsx")
            wb.save(file)
            wb.close()
            main.main()

    root = Tk()
    root.title("Registration")
    reg_label = Label(root, text="Registration!", bg="black", fg="white", width=20, font=("bold", 20))
    reg_label.place(x=90, y=53)
    reg_label.pack()
    width, height = 500, 500
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))
    root.geometry("{}x{}+{}+{}".format(width, height, x, y))

    register_name = Entry(root)
    register_email = Entry(root)
    currentdir = os.getcwd()
    file = os.path.join(currentdir, "excel.xlsx")
    wb = load_workbook(file)
    sheet = wb.active

    usrn_label = Label(root, text="Username:", bg="black", fg="white", width=18, font=("bold", 10))
    usrn_label.place(x=68, y=130)

    register_name.place(x=240, y=130)
    register_name.bind("<Return>", regname_focus)

    email_label = Label(root, text="Email:", bg="black", fg="white", width=18, font=("bold", 10))
    email_label.place(x=68, y=180)

    register_email.place(x=240, y=180)
    register_email.bind("<Return>", regemail_focus)

    register_logintype = Label(root, text="Login type:", fg="white", bg="black", width=18, font=("bold", 10))
    register_logintype.place(x=70, y=230)

    var = IntVar()

    Radiobutton(root, text="Words", padx=5, variable=var, value=1).place(x=235, y=230)

    Radiobutton(root, text="Images", padx=5, variable=var, value=2).place(x=290, y=230)

    register_logintype.bind("<Return>", reglogintype_focus)

    Button(root, text="Submit", width=20, bg="purple", fg="white", font="Helvetica", relief=SOLID,
           command=wordsorimages).place(x=150, y=300)

    Button(root, text="Go back", width=20, bg="purple", fg="white", font="Helvetica", relief=SOLID,
           command=prevpage).place(x=150, y=360)

    excel()

    root.configure(bg='black')

    root.mainloop()


if __name__ == "__main__":
    reg()
